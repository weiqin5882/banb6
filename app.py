from __future__ import annotations

import io
import re
import uuid
from datetime import datetime
from typing import Dict, Tuple

import pandas as pd
from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

app = Flask(__name__)

REPORT_CACHE: Dict[str, Dict] = {}
ALLOWED_STATUS = {"交易成功", "已发货", "已收货"}


def read_uploaded_excel(file_storage) -> pd.DataFrame:
    filename = (file_storage.filename or "").lower()
    if filename.endswith(".et"):
        raise ValueError("暂不支持 .et 直接解析，请在 WPS 中另存为 .xlsx 后再上传。")

    raw = file_storage.read()
    if not raw:
        raise ValueError("上传文件为空。")

    try:
        df = pd.read_excel(io.BytesIO(raw), dtype=str)
    except Exception as exc:
        raise ValueError(f"Excel 读取失败：{exc}") from exc

    if df.empty:
        raise ValueError("表格没有可用数据。")

    df.columns = [str(c).strip() for c in df.columns]
    return df.fillna("")


def normalize_order_no(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    text = re.sub(r"\s+", "", text)
    return re.sub(r"\D", "", text)


def to_number(value, default: float = 0.0) -> float:
    text = str(value or "").strip()
    if text == "":
        return float(default)
    text = re.sub(r"[¥,\s]", "", text)
    try:
        return float(text)
    except ValueError:
        return float(default)


def standardize(
    df: pd.DataFrame,
    mapping: dict,
    source_name: str,
    default_cost: float,
    filter_status: bool,
) -> Tuple[pd.DataFrame, dict]:
    required = ["order_no", "product_name", "sales_amount"]
    missing_required = [k for k in required if not mapping.get(k)]
    if missing_required:
        raise ValueError(f"{source_name} 映射缺失字段：{', '.join(missing_required)}")

    for key, col in mapping.items():
        if col and col not in df.columns:
            raise ValueError(f"{source_name} 映射字段不存在：{key} -> {col}")

    work = pd.DataFrame()
    work["order_no"] = df[mapping["order_no"]].map(normalize_order_no)
    work["product_name"] = df[mapping["product_name"]].astype(str).str.strip()
    work["sales_amount"] = df[mapping["sales_amount"]].map(lambda x: to_number(x, 0))
    cost_col = mapping.get("cost_amount")
    if cost_col:
        work["cost_amount"] = df[cost_col].map(lambda x: to_number(x, default_cost))
    else:
        work["cost_amount"] = float(default_cost)

    status_col = mapping.get("status")
    if status_col:
        work["status"] = df[status_col].astype(str).str.strip()
    else:
        work["status"] = ""

    total_before = len(work)

    work = work[work["order_no"] != ""].copy()
    empty_removed = total_before - len(work)

    duplicate_rows = int(work.duplicated(subset=["order_no"], keep=False).sum())
    work = work.drop_duplicates(subset=["order_no"], keep="first").copy()

    filtered_out = 0
    if filter_status and status_col:
        before_filter = len(work)
        work = work[work["status"].isin(ALLOWED_STATUS)].copy()
        filtered_out = before_filter - len(work)

    stats = {
        "source": source_name,
        "total_rows": total_before,
        "empty_order_removed": empty_removed,
        "duplicate_rows": duplicate_rows,
        "status_filtered_rows": filtered_out,
        "kept_rows": len(work),
    }

    return work, stats


def compare_orders(official: pd.DataFrame, service: pd.DataFrame) -> pd.DataFrame:
    official_map = official.set_index("order_no", drop=False)
    service_map = service.set_index("order_no", drop=False)

    official_set = set(official_map.index)
    service_set = set(service_map.index)

    matched_ids = sorted(official_set & service_set)
    official_missing = sorted(official_set - service_set)
    service_extra = sorted(service_set - official_set)

    rows = []

    for order_no in matched_ids:
        o = official_map.loc[order_no]
        s = service_map.loc[order_no]
        sales = o["sales_amount"] if o["sales_amount"] != 0 else s["sales_amount"]
        cost = s["cost_amount"] if s["cost_amount"] != 0 else o["cost_amount"]
        profit = round(float(sales) - float(cost), 2)
        tx_status = o["status"] or s["status"]
        rows.append(
            {
                "订单号": order_no,
                "商品名称": s["product_name"] or o["product_name"],
                "销售额": round(float(sales), 2),
                "成本": round(float(cost), 2),
                "利润": profit,
                "状态": tx_status,
                "比对结果": "匹配",
            }
        )

    for order_no in official_missing:
        o = official_map.loc[order_no]
        profit = round(float(o["sales_amount"]) - float(o["cost_amount"]), 2)
        rows.append(
            {
                "订单号": order_no,
                "商品名称": o["product_name"],
                "销售额": round(float(o["sales_amount"]), 2),
                "成本": round(float(o["cost_amount"]), 2),
                "利润": profit,
                "状态": o["status"],
                "比对结果": "客服漏记",
            }
        )

    for order_no in service_extra:
        s = service_map.loc[order_no]
        profit = round(float(s["sales_amount"]) - float(s["cost_amount"]), 2)
        rows.append(
            {
                "订单号": order_no,
                "商品名称": s["product_name"],
                "销售额": round(float(s["sales_amount"]), 2),
                "成本": round(float(s["cost_amount"]), 2),
                "利润": profit,
                "状态": s["status"],
                "比对结果": "异常订单",
            }
        )

    result = pd.DataFrame(rows)
    if result.empty:
        return pd.DataFrame(columns=["类序号", "订单号", "商品名称", "销售额", "成本", "利润", "状态", "比对结果"])

    compare_order = {"匹配": 0, "客服漏记": 1, "异常订单": 2}
    result["_rank"] = result["比对结果"].map(compare_order)
    result = result.sort_values(by=["_rank", "订单号"]).drop(columns=["_rank"]).reset_index(drop=True)
    result.insert(0, "类序号", range(1, len(result) + 1))
    return result


def build_summary(result_df: pd.DataFrame, official_stats: dict, service_stats: dict) -> dict:
    summary_status = {"交易成功", "已收货"}
    matched_df = result_df[result_df["比对结果"] == "匹配"].copy() if not result_df.empty else result_df
    matched_summary_df = (
        matched_df[matched_df["状态"].isin(summary_status)].copy()
        if not matched_df.empty
        else matched_df
    )

    total_sales = float(matched_summary_df["销售额"].sum()) if not matched_summary_df.empty else 0.0
    total_cost = float(matched_summary_df["成本"].sum()) if not matched_summary_df.empty else 0.0
    total_profit = float(matched_summary_df["利润"].sum()) if not matched_summary_df.empty else 0.0

    matched_count = int(len(matched_df)) if not matched_df.empty else 0
    summary_count = int(len(matched_summary_df)) if not matched_summary_df.empty else 0
    missing_count = int((result_df["比对结果"] == "客服漏记").sum()) if not result_df.empty else 0
    abnormal_count = int((result_df["比对结果"] == "异常订单").sum()) if not result_df.empty else 0
    loss_count = int((matched_summary_df["利润"] < 0).sum()) if not matched_summary_df.empty else 0

    return {
        "total_sales": round(total_sales, 2),
        "total_cost": round(total_cost, 2),
        "total_profit": round(total_profit, 2),
        "order_count": int(len(result_df)),
        "matched_count": matched_count,
        "summary_count": summary_count,
        "missing_count": missing_count,
        "abnormal_count": abnormal_count,
        "loss_count": loss_count,
        "official_stats": official_stats,
        "service_stats": service_stats,
    }


def export_excel_bytes(result_df: pd.DataFrame, summary: dict) -> io.BytesIO:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        result_df.to_excel(writer, sheet_name="对账结果", index=False)

    output.seek(0)
    wb = load_workbook(output)
    ws = wb["对账结果"]

    blue_fill = PatternFill(fill_type="solid", fgColor="2F75B5")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    red_font = Font(color="FF0000", bold=True)

    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = blue_fill

    headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    col_profit = headers.get("利润")
    col_compare = headers.get("比对结果")

    for row_idx in range(2, ws.max_row + 1):
        if col_profit:
            profit_value = ws.cell(row=row_idx, column=col_profit).value
            if isinstance(profit_value, (int, float)) and profit_value < 0:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).font = red_font

        if col_compare:
            compare_value = ws.cell(row=row_idx, column=col_compare).value
            if compare_value in {"客服漏记", "异常订单"}:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = yellow_fill

    summary_start = ws.max_row + 2
    ws.cell(row=summary_start, column=1, value="汇总统计")
    ws.cell(row=summary_start, column=1).font = Font(bold=True)

    summary_rows = [
        ("总销售额", summary["total_sales"]),
        ("总成本", summary["total_cost"]),
        ("总利润", summary["total_profit"]),
        ("订单总数", summary["order_count"]),
        ("客服漏记", summary["missing_count"]),
        ("异常订单", summary["abnormal_count"]),
        ("亏损订单", summary["loss_count"]),
    ]
    for i, (k, v) in enumerate(summary_rows, start=1):
        ws.cell(row=summary_start + i, column=1, value=k)
        ws.cell(row=summary_start + i, column=2, value=v)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(12, max_len + 2), 40)

    final_out = io.BytesIO()
    wb.save(final_out)
    final_out.seek(0)
    return final_out


@app.route("/")
def index():
    return render_template("index.html")


@app.post("/api/inspect")
def inspect_file():
    try:
        file = request.files.get("file")
        if not file:
            return jsonify({"ok": False, "message": "请上传文件。"}), 400

        df = read_uploaded_excel(file)
        preview = df.head(3).to_dict(orient="records")
        return jsonify({
            "ok": True,
            "columns": list(df.columns),
            "rows": len(df),
            "preview": preview,
        })
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400


@app.post("/api/compare")
def compare_api():
    try:
        official_file = request.files.get("official_file")
        service_file = request.files.get("service_file")
        if not official_file or not service_file:
            return jsonify({"ok": False, "message": "请同时上传官方和客服文件。"}), 400

        official_mapping = {
            "order_no": request.form.get("official_order_no", ""),
            "status": request.form.get("official_status", ""),
            "product_name": request.form.get("official_product_name", ""),
            "sales_amount": request.form.get("official_sales_amount", ""),
            "cost_amount": request.form.get("official_cost_amount", ""),
        }
        service_mapping = {
            "order_no": request.form.get("service_order_no", ""),
            "status": request.form.get("service_status", ""),
            "product_name": request.form.get("service_product_name", ""),
            "sales_amount": request.form.get("service_sales_amount", ""),
            "cost_amount": request.form.get("service_cost_amount", ""),
        }

        default_cost = to_number(request.form.get("default_cost", "0"), 0.0)

        if not official_mapping["status"]:
            return jsonify({"ok": False, "message": "官方订单必须映射“交易状态”字段。"}), 400

        official_df = read_uploaded_excel(official_file)
        service_df = read_uploaded_excel(service_file)

        official_clean, official_stats = standardize(
            official_df, official_mapping, "官方订单", default_cost, filter_status=True
        )
        service_clean, service_stats = standardize(
            service_df, service_mapping, "客服订单", default_cost, filter_status=False
        )

        result_df = compare_orders(official_clean, service_clean)
        summary = build_summary(result_df, official_stats, service_stats)

        report_id = uuid.uuid4().hex
        REPORT_CACHE[report_id] = {
            "result_df": result_df,
            "summary": summary,
            "created_at": datetime.now().isoformat(),
        }

        return jsonify({
            "ok": True,
            "report_id": report_id,
            "summary": summary,
            "total_rows": len(result_df),
        })
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400
    except Exception as exc:
        return jsonify({"ok": False, "message": f"处理失败：{exc}"}), 500


@app.get("/api/report/<report_id>")
def report_page(report_id: str):
    data = REPORT_CACHE.get(report_id)
    if not data:
        return jsonify({"ok": False, "message": "报告不存在或已过期，请重新处理。"}), 404

    page = max(int(request.args.get("page", 1)), 1)
    page_size = min(max(int(request.args.get("page_size", 50)), 10), 500)

    result_df = data["result_df"]
    total_rows = len(result_df)
    total_pages = max((total_rows + page_size - 1) // page_size, 1)

    if page > total_pages:
        page = total_pages

    start = (page - 1) * page_size
    end = start + page_size

    records = result_df.iloc[start:end].to_dict(orient="records") if total_rows else []

    return jsonify({
        "ok": True,
        "records": records,
        "page": page,
        "page_size": page_size,
        "total_rows": total_rows,
        "total_pages": total_pages,
        "summary": data["summary"],
    })


@app.get("/api/export/<report_id>")
def export_report(report_id: str):
    data = REPORT_CACHE.get(report_id)
    if not data:
        return jsonify({"ok": False, "message": "报告不存在或已过期，请重新处理。"}), 404

    excel_io = export_excel_bytes(data["result_df"], data["summary"])
    today = datetime.now().strftime("%Y%m%d")
    return send_file(
        excel_io,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"订单对账结果_{today}.xlsx",
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8000, debug=True)
